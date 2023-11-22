import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys
curPath = ""
if getattr(sys, 'frozen', False):
    # If the application is run as a bundle (e.g. PyInstaller bundle)
    os.chdir(sys._MEIPASS)
else:
    # If the application is run as a script
    curPath = os.path.join(os.getcwd(),'cognosWorker')
    pass

def export_to_excel(filename,outputPath,modelPath_df,df_query_dataItems,df_promptPages,df_reportPages,df_pageNames):
  exportFileName = os.path.join(outputPath,filename[:-4]+".xlsx")
  with pd.ExcelWriter(exportFileName) as writer:
    
      # use to_excel function and specify the sheet_name and index
      # to store the dataframe in specified sheet
      # modelPath_df.to_excel(writer, sheet_name="Model Path", index=False)
      df_query_dataItems.to_excel(writer, sheet_name="Query Details", index=False)
      df_pageNames.to_excel(writer, sheet_name="Pages", index=False)
      df_promptPages.to_excel(writer, sheet_name="Prompt Page", index=False)
      df_reportPages.to_excel(writer, sheet_name="Report Pages", index=False)
  print(f"[+] {filename[:-4]}.xlsx file created")


# def save_complexity_report(complexity_df,complexity_sheet_name,complexity_workbook_name,outputPath):
#     complexity_workbook_name = os.path.join(curPath,complexity_workbook_name)
#     workbook = openpyxl.load_workbook(complexity_workbook_name, keep_vba=True)
#     if complexity_sheet_name not in workbook.sheetnames:
#         workbook.create_sheet(complexity_sheet_name)
#
#     worksheet = workbook[complexity_sheet_name]
#     worksheet.delete_rows(1, worksheet.max_row)
#     rows = dataframe_to_rows(complexity_df, index=False, header=True)
#     for r_idx, row in enumerate(rows, 1):
#         for c_idx, value in enumerate(row, 1):
#             worksheet.cell(row=r_idx, column=c_idx, value=value)
#     workbook.save(os.path.join(outputPath, r'COGNOS Complexity Report.xlsm'))
#
#
# def save_similarity_in_Complexity_Sheet(outputPath,mean_df,df_data_expression_similarity,df_dataObjects_similarity,df_dataItem_names_similarity,df_reportstaticElement_similarity,df_reportdataElement_similarity):
#   complexity_workbook_name=os.path.join(outputPath,r'COGNOS Complexity Report.xlsm')
#   print(complexity_workbook_name)
#   # print('aaaaaaaaaasdddddddasdasdasdasdasdasdssssssssssssssssssssssssssssssssssssssssssssssssss')
#   workbook = openpyxl.load_workbook(complexity_workbook_name, keep_vba=True)
#
#   complexity_sheet_name = "Report_Similarity_Score"
#   if complexity_sheet_name not in workbook.sheetnames:
#     workbook.create_sheet(complexity_sheet_name)
#   worksheet = workbook[complexity_sheet_name]
#   worksheet.delete_rows(1, worksheet.max_row)
#   rows = dataframe_to_rows(mean_df, index=False, header=True)
#   for r_idx, row in enumerate(rows, 1):
#       for c_idx, value in enumerate(row, 1):
#           worksheet.cell(row=r_idx, column=c_idx, value=value)
#
#   complexity_sheet_name = "DataExpr_Similarity_Score"
#   if complexity_sheet_name not in workbook.sheetnames:
#     workbook.create_sheet(complexity_sheet_name)
#   worksheet = workbook[complexity_sheet_name]
#   worksheet.delete_rows(1, worksheet.max_row)
#   rows = dataframe_to_rows(df_data_expression_similarity, index=False, header=True)
#   for r_idx, row in enumerate(rows, 1):
#       for c_idx, value in enumerate(row, 1):
#           worksheet.cell(row=r_idx, column=c_idx, value=value)
#
#   complexity_sheet_name = "DataObject_Similarity_Score"
#   if complexity_sheet_name not in workbook.sheetnames:
#     workbook.create_sheet(complexity_sheet_name)
#   worksheet = workbook[complexity_sheet_name]
#   worksheet.delete_rows(1, worksheet.max_row)
#   rows = dataframe_to_rows(df_dataObjects_similarity, index=False, header=True)
#   for r_idx, row in enumerate(rows, 1):
#       for c_idx, value in enumerate(row, 1):
#           worksheet.cell(row=r_idx, column=c_idx, value=value)
#
#   complexity_sheet_name = "DataItemName_Similarity_Score"
#   if complexity_sheet_name not in workbook.sheetnames:
#     workbook.create_sheet(complexity_sheet_name)
#   worksheet = workbook[complexity_sheet_name]
#   worksheet.delete_rows(1, worksheet.max_row)
#   rows = dataframe_to_rows(df_dataItem_names_similarity, index=False, header=True)
#   for r_idx, row in enumerate(rows, 1):
#       for c_idx, value in enumerate(row, 1):
#           worksheet.cell(row=r_idx, column=c_idx, value=value)
#
#   complexity_sheet_name = "InreportStaticElems_Sim_Score"
#   if complexity_sheet_name not in workbook.sheetnames:
#     workbook.create_sheet(complexity_sheet_name)
#   worksheet = workbook[complexity_sheet_name]
#   worksheet.delete_rows(1, worksheet.max_row)
#   rows = dataframe_to_rows(df_reportstaticElement_similarity, index=False, header=True)
#   for r_idx, row in enumerate(rows, 1):
#       for c_idx, value in enumerate(row, 1):
#           worksheet.cell(row=r_idx, column=c_idx, value=value)
#
#   complexity_sheet_name = "InreportDataElems_Sim_Score"
#   if complexity_sheet_name not in workbook.sheetnames:
#     workbook.create_sheet(complexity_sheet_name)
#   worksheet = workbook[complexity_sheet_name]
#   worksheet.delete_rows(1, worksheet.max_row)
#   rows = dataframe_to_rows(df_reportdataElement_similarity, index=False, header=True)
#   for r_idx, row in enumerate(rows, 1):
#       for c_idx, value in enumerate(row, 1):
#           worksheet.cell(row=r_idx, column=c_idx, value=value)
#
#   workbook.save(complexity_workbook_name)